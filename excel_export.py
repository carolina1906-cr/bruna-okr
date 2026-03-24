import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from db import get_departments, get_key_results, get_monthly_values
from calculations import calcular_avance, semaforo
from constants import MESES, COLORES

FILL = {
    "sobre_meta": PatternFill("solid", fgColor="1A2744"),
    "en_meta":    PatternFill("solid", fgColor="2DC653"),
    "en_riesgo":  PatternFill("solid", fgColor="FFD600"),
    "critico":    PatternFill("solid", fgColor="E63946"),
    "sin_dato":   PatternFill("solid", fgColor="D0D4DF"),
}
FONT_BLANCO = Font(color="FFFFFF", bold=True)
FONT_OSCURO = Font(color="1A2744", bold=True)

def estilo_semaforo(ws, cell, estado):
    cell.fill = FILL.get(estado, FILL["sin_dato"])
    if estado in ("sobre_meta", "en_meta", "critico"):
        cell.font = FONT_BLANCO
    else:
        cell.font = FONT_OSCURO
    cell.alignment = Alignment(horizontal="center")

def generar_excel(year, mes_activo):
    wb = Workbook()
    departments = get_departments()
    all_krs = get_key_results()

    # Hoja Dashboard
    ws = wb.active
    ws.title = "Dashboard"
    headers = ["Departamento", "KR", "Unidad", "Meta", "% Mes", "% Acum.", "Estado"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1A2744")
        cell.alignment = Alignment(horizontal="center")

    for dept in departments:
        krs = [k for k in all_krs if k["department_code"] == dept["code"]]
        for kr in krs:
            vals = get_monthly_values(kr["id"], year)
            pct_m, pct_a = calcular_avance(kr, vals, mes_activo)
            estado = semaforo(pct_m)
            row = [
                dept["name"], kr["name"], kr["unit"], kr["goal"],
                round(pct_m, 1) if pct_m is not None else None,
                round(pct_a, 1) if pct_a is not None else None,
                estado.replace("_", " ").title()
            ]
            ws.append(row)
            r = ws.max_row
            estilo_semaforo(ws, ws.cell(r, 5), estado)
            estilo_semaforo(ws, ws.cell(r, 6), estado)

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["G"].width = 14

    # Una hoja por departamento
    for dept in departments:
        ws2 = wb.create_sheet(dept["name"][:31])
        ws2.append(["KR", "Unidad", "Base", "Meta"] + MESES + ["% Mes", "% Acum."])
        for cell in ws2[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1A2744")

        krs = [k for k in all_krs if k["department_code"] == dept["code"]]
        for kr in krs:
            vals = get_monthly_values(kr["id"], year)
            pct_m, pct_a = calcular_avance(kr, vals, mes_activo)
            estado = semaforo(pct_m)
            row = [kr["name"], kr["unit"], kr.get("base"), kr["goal"]]
            row += [vals.get(m) for m in range(1, 13)]
            row += [
                round(pct_m, 1) if pct_m is not None else None,
                round(pct_a, 1) if pct_a is not None else None,
            ]
            ws2.append(row)
            r = ws2.max_row
            estilo_semaforo(ws2, ws2.cell(r, 17), estado)
            estilo_semaforo(ws2, ws2.cell(r, 18), estado)
        ws2.column_dimensions["A"].width = 45

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
